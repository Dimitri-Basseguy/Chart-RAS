#!/usr/bin/env python3
"""
import_agences.py — Importe la feuille "STATS Agence" d'un fichier Excel → agences.js

Usage :
    python3 import_agences.py lead-mars-2026.xlsx   # ajoute/met à jour ce mois
    python3 import_agences.py                        # relit tous les lead-*.xlsx (mode legacy)
"""

import openpyxl, json, glob, re, os, sys

SHEET   = 'STATS Agence'
FILE_RE = re.compile(r'lead-(\w+)-(\d{4})\.xlsx', re.IGNORECASE)

MONTH_NAMES = {
    'janv':'Janvier','fev':'Février','mars':'Mars','avr':'Avril',
    'mai':'Mai','juin':'Juin','juil':'Juillet','aout':'Août',
    'sept':'Septembre','oct':'Octobre','nov':'Novembre','dec':'Décembre',
}
MONTH_NUMS = {
    'janv':'01','fev':'02','mars':'03','avr':'04','mai':'05','juin':'06',
    'juil':'07','aout':'08','sept':'09','oct':'10','nov':'11','dec':'12',
}

COL_MAP = {
    'rasagence_email_cv':    'agence',
    'cv reçus':              'nb_cv',
    '% total mesurable':     'pct_total',
    'intérimaires':          'nb_int',
    'nbres new int':         'n_int',
    '% new int':             'pct_n_int',
    "tx de mise à l'emploi":'tx_emploi',
    'ca hrfa':               'ca',
    'marge':                 'marge',
    'régions':               'region',
    'roi marge':             'roi_marge',
    'cout prorata cv':       'cout_prorata',
}


def parse_filename(filename):
    m = FILE_RE.match(os.path.basename(filename))
    if not m:
        return None, None
    mois = m.group(1).lower()
    annee = m.group(2)
    num  = MONTH_NUMS.get(mois)
    name = MONTH_NAMES.get(mois)
    if not num:
        print(f"  ⚠  Mois non reconnu : {mois}")
        return None, None
    return f"{annee}-{num}", f"{name} {annee}"


def import_sheet(path):
    wb   = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        print(f"  ⚠  Feuille '{SHEET}' absente de {os.path.basename(path)}")
        return []

    rows   = list(wb[SHEET].iter_rows(values_only=True))
    header = [str(h).strip().lower() if h else '' for h in rows[0]]

    idx = {COL_MAP[h]: i for i, h in enumerate(header) if h in COL_MAP}
    if 'agence' not in idx:
        print(f"  ⚠  Colonne agence introuvable dans {os.path.basename(path)}")
        return []

    result = []
    for row in rows[1:]:
        agence = row[idx['agence']]
        if not agence:
            continue

        def get(field):
            if field not in idx:
                return None
            val = row[idx[field]]
            if val is None:
                return None
            if field == 'region':
                return str(val).strip()
            try:
                return round(float(val), 2)
            except (TypeError, ValueError):
                return None

        result.append({
            'agence':       str(agence).strip(),
            'nb_cv':        get('nb_cv'),
            'pct_total':    get('pct_total'),
            'nb_int':       get('nb_int'),
            'n_int':        get('n_int'),
            'pct_n_int':    get('pct_n_int'),
            'tx_emploi':    get('tx_emploi'),
            'ca':           get('ca'),
            'marge':        get('marge'),
            'region':       get('region'),
            'roi_marge':    get('roi_marge'),
            'cout_prorata': get('cout_prorata'),
        })

    return result


def load_existing():
    if not os.path.exists('agences.js'):
        return {}
    try:
        with open('agences.js', encoding='utf-8') as f:
            content = f.read()
        js_obj = content.replace('const AGENCES_DATA =', '', 1).rstrip().rstrip(';')
        return json.loads(js_obj)
    except Exception:
        return {}


def save(data):
    js = 'const AGENCES_DATA = ' + json.dumps(data, ensure_ascii=False, indent=2) + ';\n'
    with open('agences.js', 'w', encoding='utf-8') as f:
        f.write(js)
    total = sum(len(v['rows']) for v in data.values())
    print(f"\n✓ agences.js — {len(data)} mois, {total} agences, {len(js):,} car.")
    print("  Étape suivante : python3 encrypt_agences.py")


def import_one(excel_path):
    key, label = parse_filename(excel_path)
    if not key:
        print(f"✗ Impossible de détecter le mois depuis '{excel_path}'.")
        print("  Nommez le fichier : lead-mars-2026.xlsx")
        sys.exit(1)

    print(f"  {excel_path}  →  {key} ({label})")
    rows = import_sheet(excel_path)
    if not rows:
        print("✗ Aucune agence trouvée.")
        sys.exit(1)
    print(f"    {len(rows)} agences")

    existing = load_existing()
    existing[key] = {'label': label, 'rows': rows}
    save(dict(sorted(existing.items())))


def import_all():
    files = sorted(glob.glob('lead-*.xlsx'))
    if not files:
        print("✗ Aucun fichier lead-*.xlsx trouvé.")
        sys.exit(1)

    data = {}
    for f in files:
        key, label = parse_filename(f)
        if not key:
            continue
        print(f"  {f}  →  {key} ({label})")
        rows = import_sheet(f)
        print(f"    {len(rows)} agences")
        data[key] = {'label': label, 'rows': rows}

    if not data:
        print("✗ Aucune donnée importée.")
        sys.exit(1)

    save(data)


if __name__ == '__main__':
    if len(sys.argv) >= 2:
        if not os.path.exists(sys.argv[1]):
            print(f"✗ Fichier introuvable : {sys.argv[1]}")
            sys.exit(1)
        import_one(sys.argv[1])
    else:
        import_all()
