#!/usr/bin/env python3
"""
import_regions.py — Importe la feuille "Régions" d'un fichier Excel → data_regions.js

Usage :
    python3 import_regions.py lead-mars-2026.xlsx   # ajoute/met à jour ce mois
    python3 import_regions.py                        # relit tous les lead-*.xlsx (mode legacy)
"""

import openpyxl, json, glob, re, os, sys

SHEET         = 'Régions'
FILE_RE       = re.compile(r'lead-(\w+)-(\d{4})\.xlsx', re.IGNORECASE)
REGION_MARKER = 'choix région =>'
MAX_DATA_ROWS = 19
SKIP_SOURCES  = ["résumé", "resume", "total", "calcul"]
MANUAL_FIELDS = ['offres']

MONTH_NAMES = {
    'janv':'Janvier','fev':'Février','mars':'Mars','avr':'Avril',
    'mai':'Mai','juin':'Juin','juil':'Juillet','aout':'Août',
    'sept':'Septembre','oct':'Octobre','nov':'Novembre','dec':'Décembre',
}
MONTH_NUMS = {
    'janv':'01','fev':'02','mars':'03','avr':'04','mai':'05','juin':'06',
    'juil':'07','aout':'08','sept':'09','oct':'10','nov':'11','dec':'12',
}

COL_MAP = {
    'sources':                'source',
    'budget':                 'budget',
    'nb cv':                  'nb_cv',
    'cpa cv':                 'cpa_cv',
    'candidat uniq':          'candidat_uniq',
    'new candidats validés':  'new_cand',
    'tx new cand validés':    'tx_new_cand',
    'cpnc':                   'cpnc',
    'nb intérimaires':        'nb_int',
    "tx mise à l'emploi":     'tx_emploi',
    'cme':                    'cme',
    'nint':                   'n_int',
    '% nint':                 'pct_n_int',
    'ca total hrfa':          'ca',
    'roi brut ca':            'roi_brut',
    'marge total':            'marge',
    '% marge':                'pct_marge',
    'roi réel marge':         'roi_reel',
}

GRATUIT_SOURCES = {
    "site carrière", "cvthèque", "candidature spontanée",
    "gmb", "google my business", "google jobs", "google jobs apply",
    "apec", "jooble", "jooble.org", "france travail", "talent", "talent.com",
    "linkedin limited", "monster organic", "truckfly",
}


def parse_filename(filename):
    m = FILE_RE.match(os.path.basename(filename))
    if not m:
        return None, None
    mois = m.group(1).lower()
    annee = m.group(2)
    num  = MONTH_NUMS.get(mois)
    name = MONTH_NAMES.get(mois)
    if not num:
        print(f"  ⚠  Mois non reconnu : {mois}")
        return None, None
    return f"{annee}-{num}", f"{name} {annee}"


def clean_val(v):
    if v is None:
        return None
    s = str(v).strip()
    if s in ('', '-', '—', '#DIV/0!', 'nan', 'NaN', 'N/A', '#N/A', 'None'):
        return None
    s = s.replace('\xa0', '').replace(' ', '').replace('€', '').replace(',', '.')
    if s.endswith('%'):
        try:
            return round(float(s[:-1]) / 100, 6)
        except ValueError:
            return None
    try:
        return round(float(s), 2)
    except (ValueError, TypeError):
        return None


def detect_type(source_name, budget):
    if source_name.lower().strip() in GRATUIT_SOURCES:
        return "gratuit"
    try:
        if budget and float(budget) > 0:
            return "payant"
    except (TypeError, ValueError):
        pass
    return "gratuit"


def import_regions_sheet(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        print(f"  ⚠  Feuille '{SHEET}' absente de {os.path.basename(path)}")
        return {}

    all_rows = list(wb[SHEET].iter_rows(values_only=True))
    regions  = {}
    i = 0

    while i < len(all_rows):
        row   = all_rows[i]
        first = str(row[0]).strip().lower() if row[0] else ''

        if first == REGION_MARKER:
            region_name = str(row[1]).strip() if row[1] else None
            if not region_name or region_name.lower() in ('none', ''):
                i += 1
                continue

            i += 1
            if i >= len(all_rows):
                break
            header = [str(h).strip().lower() if h else '' for h in all_rows[i]]

            col_idx = {}
            for j, h in enumerate(header):
                for col_key, json_key in COL_MAP.items():
                    if json_key not in col_idx.values() and (col_key in h or h in col_key):
                        col_idx[j] = json_key
                        break

            data_rows = []
            i += 1
            count = 0
            while i < len(all_rows) and count < MAX_DATA_ROWS:
                r = all_rows[i]
                src = str(r[0]).strip() if r[0] is not None else ''

                if not src or src.lower() in ('none', ''):
                    break
                if any(kw in src.lower() for kw in SKIP_SOURCES):
                    i += 1
                    continue

                obj = {'source': src}
                for cidx, jkey in col_idx.items():
                    if jkey == 'source':
                        continue
                    obj[jkey] = clean_val(r[cidx]) if cidx < len(r) else None

                for field in ['budget','nb_cv','cpa_cv','candidat_uniq','new_cand',
                              'tx_new_cand','cpnc','nb_int','tx_emploi','cme',
                              'n_int','pct_n_int','ca','roi_brut','marge','pct_marge','roi_reel']:
                    obj.setdefault(field, None)

                budget = obj.get('budget') or 0
                obj['budget'] = budget if budget else 0
                obj['type'] = detect_type(src, budget)

                data_rows.append(obj)
                count += 1
                i += 1

            regions[region_name] = {'rows': data_rows}
            print(f"    '{region_name}' : {len(data_rows)} sources")
        else:
            i += 1

    return regions


def load_existing():
    """Charge data_regions.js et retourne le dict existant."""
    if not os.path.exists('data_regions.js'):
        return {}
    try:
        with open('data_regions.js', encoding='utf-8') as f:
            content = f.read()
        js_obj = content.replace('const REGIONS_DATA =', '', 1).rstrip().rstrip(';')
        return json.loads(js_obj)
    except Exception:
        return {}


def preserve_manual(regions, existing_month):
    """Recopie les champs manuels (offres) depuis l'existant vers les nouvelles données."""
    for region_name, region_data in regions.items():
        for field in MANUAL_FIELDS:
            existing_val = existing_month.get('regions', {}).get(region_name, {}).get(field)
            if existing_val is not None:
                region_data[field] = existing_val
                print(f"    Champ '{field}' préservé pour {region_name}")


def save(data):
    js = 'const REGIONS_DATA = ' + json.dumps(data, ensure_ascii=False, indent=2) + ';\n'
    with open('data_regions.js', 'w', encoding='utf-8') as f:
        f.write(js)
    total = sum(len(v['regions']) for v in data.values())
    print(f"\n✓ data_regions.js — {len(data)} mois, {total} blocs région, {len(js):,} car.")
    print("  Étape suivante : python3 encrypt_regions.py")


def import_one(excel_path):
    """Importe un seul fichier et met à jour data_regions.js."""
    key, label = parse_filename(excel_path)
    if not key:
        print(f"✗ Impossible de détecter le mois depuis '{excel_path}'.")
        print("  Nommez le fichier : lead-mars-2026.xlsx")
        sys.exit(1)

    print(f"  {excel_path}  →  {key} ({label})")
    regions = import_regions_sheet(excel_path)
    if not regions:
        print("✗ Aucune région trouvée.")
        sys.exit(1)

    existing = load_existing()
    preserve_manual(regions, existing.get(key, {}))

    existing[key] = {'label': label, 'regions': regions}
    # Trier les mois par ordre chronologique
    data = dict(sorted(existing.items()))
    save(data)


def import_all():
    """Relit tous les lead-*.xlsx (mode legacy)."""
    files = sorted(glob.glob('lead-*.xlsx'))
    if not files:
        print("✗ Aucun fichier lead-*.xlsx trouvé.")
        sys.exit(1)

    existing = load_existing()
    data = {}
    for f in files:
        key, label = parse_filename(f)
        if not key:
            continue
        print(f"  {f}  →  {key} ({label})")
        regions = import_regions_sheet(f)
        if not regions:
            print(f"    ⚠  Aucune région trouvée")
            continue
        preserve_manual(regions, existing.get(key, {}))
        data[key] = {'label': label, 'regions': regions}

    if not data:
        print("✗ Aucune donnée importée.")
        sys.exit(1)

    save(data)


if __name__ == '__main__':
    if len(sys.argv) >= 2:
        if not os.path.exists(sys.argv[1]):
            print(f"✗ Fichier introuvable : {sys.argv[1]}")
            sys.exit(1)
        import_one(sys.argv[1])
    else:
        import_all()
